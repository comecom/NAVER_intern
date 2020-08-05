from openpyxl import load_workbook, Workbook

names = ['문재인', '전현무', '한혜진', '강다니엘', '방탄', '소년', '뷔', '손흥민', '이찬원', '지민',
         '탁재훈', '이가흔', '이영탁', '김호중', '세븐', '김강열', '송가인', '추미애',
         '이숭우', '정국', '은지원', '유희열', '엑스원', '김요한', '고영수', '김우석', '재석', '헨리', '사콜',
         '사랑의콜센터', '효리', '광희', '김희재', '이승기', '박원순', '윤석열', '석렬', '조인성', '박근혜',
         '최순실', '최강욱', '김정은', '노무현', '박정희', '전두환', '정동원', '신유', '찬또', '조국',
         '박지현', '트와이스', '쯔위', '모모', '사나', '정연', '나연', '태연']

load_wb = load_workbook("/Users/user/Desktop/excel_test.xlsx", data_only=True)
load_ws = load_wb['Sheet1']

write_wb = Workbook()
write_ws = write_wb.active

write_ws.cell(1, 1, 'contents')
write_ws.cell(1, 2, 'sentiment')

get_row = load_ws['A2':'B95877']

for line in get_row:
    #print(line[0].value)
    str = line[0].value
    for name in names:
        if name in str:
            str = str.replace(name, "XXX")
        if len(name) >= 3:
            first_name = name[1:]
            if first_name in str:
                str = str.replace(first_name, "XXX")
    print(line)
    write_ws.append([str, line[1].value])

for name in names:
    write_ws.append([name, 'UNK'])

write_wb.save('/Users/user/Desktop/train_rename.xlsx')