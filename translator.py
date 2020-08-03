from openpyxl import load_workbook, Workbook

load_wb = load_workbook("/Users/user/Desktop/excel_test.xlsx", data_only=True)
load_ws = load_wb['Sheet1']

write_wb = Workbook()
write_ws = write_wb.active

unk_wb = Workbook()
unk_ws = unk_wb.active

count_POS = 0
count_NEG = 0
count_UNK = 0
Blank_List = []
Error_List = []

get_row = load_ws['A2':'B95147']

for row in get_row:
    if row[1].value == 'POS':
        count_POS = count_POS+1
    elif row[1].value == 'NEG':
        count_NEG = count_NEG+1
    elif row[1].value == 'UNK':
        count_UNK = count_UNK+1
    elif row[1].value == '':
        Blank_List.append(row)
    else:
        Error_List.append(row)

print("POS : {}, NEG : {}, UNK : {}".format(count_POS, count_NEG, count_UNK))
print("total data : {}".format(count_POS+count_NEG+count_UNK))

if len(Error_List) != 0:
    for position in Error_List:
        print(position[1], position[1].value)

if len(Blank_List) != 0:
    for position in Blank_List:
        print(position[1], position[1].value)

#binary data (POS/NEG) refinement
write_ws.cell(1, 1, 'contents')
write_ws.cell(1, 2, 'sentiment')
for line in get_row:
    if line[1].value == 'UNK':
        continue
    else:
        write_ws.append([line[0].value, line[1].value])

write_wb.save('/Users/user/Desktop/train_bin.xlsx')

bin_rb = load_workbook('/Users/user/Desktop/train_bin.xlsx')
bin_sb = bin_rb['Sheet']

print("\nMULTI label -> BINARY classification")
print("translation complete!!")

row_range = bin_sb[2:count_POS+count_NEG+1]
bin_POS = 0
bin_NEG = 0
bin_Error = []
for line in row_range:
    if line[1].value == 'POS':
        bin_POS += 1
    elif line[1].value == 'NEG':
        bin_NEG += 1
    else:
        bin_Error.append(line)

print("bin_POS : {}, bin_NEG : {}".format(bin_POS, bin_NEG))

if len(bin_Error) != 0:
    for line in bin_Error:
        print(line)


#UNK data refinement
unk_ws.cell(1, 1, 'contents')
unk_ws.cell(1, 2, 'sentiment')
for line in get_row:
    if line[1].value == 'UNK':
        unk_ws.append([line[0].value, line[1].value])


unk_wb.save('/Users/user/Desktop/unk_data.xlsx')

unk_rb = load_workbook('/Users/user/Desktop/unk_data.xlsx')
unk_sb = unk_rb['Sheet']

row_range = unk_sb[2:count_UNK+1]
UNK = 0
unk_Error = []
for line in row_range:
    if line[1].value == 'UNK':
        UNK += 1
    else:
        unk_Error.append(line)

print("UNK : {}".format(UNK))

if len(unk_Error) != 0:
    for line in unk_Error:
        print(line)